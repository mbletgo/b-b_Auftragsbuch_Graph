from openpyxl import load_workbook

from py2neo.data import Node, Relationship
from py2neo import  NodeMatcher, Database, Graph


# Workbook laden
wb = load_workbook('/Users/m_b/Programmieren_Python/b+b_auftragsbuch/Teil A-D ABuch.xlsx')
#hier gehört das cvs hin


# test: die sheet namen ausgeben
#print(wb.sheetnames)

# sheet waehlen
#ws = wb["2010-2019"]

#oder aktives sheet, wenn geoeffnet
ws = wb.active

#db = Database("bolt://localhost:7687")
#graph = db.default_graph

#http://localhost:7474/db/data/
# Assuming we’ve already established a connection to the server...
#/Users/m_b/Library/Application Support/Neo4j Desktop/Application/neo4jDatabases/database-b6d06588-24fc-42fb-8d5d-fbab4380bd0c/installation-3.5.6/data
#

try:
 graph = Graph("/Users/m_b/Library/Application Support/Neo4j Desktop/Application/neo4jDatabases/database-b6d06588-24fc-42fb-8d5d-fbab4380bd0c/installation-3.5.6/data/graph.db")
except Exception:
 #print(e)
        #time_log("Couldn't connect to db! Check settings!")
exit(2)

matcher = NodeMatcher(graph)

# Liste fuer doppelte Kunden/Lieferanten
listeVorhandenKunden = []
listeVorhandenLieferanten = []

for x in range(2, 7053):
  print(x)

  datum = ws.cell(row = 1, column = x).value
  auftragsnummer = ws.cell(row = 2, column = x).value

  kunde = ws.cell(row = 3, column = x).value

  if x == 2:
    listeVorhandenKunden.append(kunde)

  # check ob vorhanden
  try:
      q = listeVorhandenKunden.index(kunde)
  except ValueError:
      q = -1

  lieferant = ws.cell(row = 4, column = x).value
  if x == 2:
    listeVorhandenLieferanten.append(lieferant)

    # check ob vorhanden
    try:
        f = listeVorhandenLieferanten.index(lieferant)
    except ValueError:
        f = -1

  # wenn kunde nicht in Liste, Knoten neu anlegen und namen ans ende der Liste
  # sonst: vorhanden Knoten benutzen
  if q == -1:
    #c = Node("Kunde", name=kunde)
    graph.cypher.execute("CREATE (k:Kunde{name:'kunde'}")
    #CREATE (n:Person { name: 'Andy', title: 'Developer' })
    listeVorhandenKunden.append(kunde)

  if f == -1:
    #d = Node("Lieferant", name=lieferant)
    graph.cypher.execute("CREATE (l:Lieferant{name:'lieferant'}")
    listeVorhandenLieferanten.append(lieferant)

  # beide knoten nciht vorhanden
  if q == -1 and f == -1:
    #cd = Relationship(c, "Auftrag", d, datum)
    graph.cypher.execute("CREATE (c)-[r:AUFTRAG]->(d)")

  # beide Knoten vorhanden
  if q != -1 and f != -1:
    #cd = Relationship(Node(matcher.match("Kunde", name=listeVorhandenKunden[q])), "Auftrag", Node(matcher.match("Kunde",name=listeVorhandenLieferanten[f])))

    eineNode = Node(matcher.match("Kunde", name=listeVorhandenKunden[q]))
    zweiteNode = Node(matcher.match("Kunde",name=listeVorhandenLieferanten[f]))

    graph.cypher.execute("CREATE (eineNode)-[r:AUFTRAG]->(zweiteNode)")



  # einer der Knoten nicht vorhanden
  if q == -1 and f != -1:
    #cd = Relationship(c, "Auftrag", Node(matcher.match("Kunde",name=listeVorhandenLieferanten[f])))

    dritteNode = Node(matcher.match("Kunde",name=listeVorhandenLieferanten[f]))
    graph.cypher.execute("CREATE (c)-[r:AUFTRAG]->(dritteNode)")


  if q != -1 and f == -1:
    #cd = Relationship(Node(matcher.match("Lieferant", name=listeVorhandenKunden[q])), "Auftrag", d)

    vierteNode = Node(matcher.match("Kunde",name=listeVorhandenLieferanten[q]))
    graph.cypher.execute("CREATE (vierteNode)-[r:AUFTRAG]->(f)")
